# -*- coding: utf-8 -*-
"""
potential.py - 场景4：高潜客户线索检索。

基于 ranking_ent_dtl_clue 企业线索表与 zq_dtl_shnews_yyy 新闻信号表，
执行条件过滤、确定性评分、推荐理由生成，并提供 Excel 导出能力。
"""

import os
import re
import uuid
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlencode

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from sqlalchemy import text

import utils.db_helper as db
from utils.file_helper import ensure_dir
from utils.mock_db import POTENTIAL_CLIENTS
from utils.rag_engine import RAGEngine
import asyncio
from openai import AsyncOpenAI
import json

DISTRICT_ALIASES = {
    "浦东": "浦东新区",
    "浦东新区": "浦东新区",
    "黄浦": "黄浦区",
    "黄浦区": "黄浦区",
    "徐汇": "徐汇区",
    "徐汇区": "徐汇区",
    "长宁": "长宁区",
    "长宁区": "长宁区",
    "静安": "静安区",
    "静安区": "静安区",
    "普陀": "普陀区",
    "普陀区": "普陀区",
    "虹口": "虹口区",
    "虹口区": "虹口区",
    "杨浦": "杨浦区",
    "杨浦区": "杨浦区",
    "闵行": "闵行区",
    "闵行区": "闵行区",
    "宝山": "宝山区",
    "宝山区": "宝山区",
    "嘉定": "嘉定区",
    "嘉定区": "嘉定区",
    "金山": "金山区",
    "金山区": "金山区",
    "松江": "松江区",
    "松江区": "松江区",
    "青浦": "青浦区",
    "青浦区": "青浦区",
    "奉贤": "奉贤区",
    "奉贤区": "奉贤区",
    "崇明": "崇明区",
    "崇明区": "崇明区",
}

INDUSTRY_KEYWORDS = [
    "现代服务业", "数字化转型", "未来产业", "先导产业", "重点支撑产业", "先进制造业",
    "人工智能", "AI", "通信", "信息技术", "软件", "互联网", "数字经济", "大数据", "云计算",
    "算力", "集成电路", "半导体", "生物医药", "医疗", "智能制造", "机器人", "新能源",
    "新材料", "汽车", "金融", "文创", "航运", "低空经济", "物联网", "工业互联网",
]

NOISE_WORDS = [
    "上海市", "上海", "推荐", "高潜", "潜在", "重点", "客户", "名单", "线索", "商机", "企业", "有哪些",
    "给我", "帮我", "筛选", "查询", "查看", "导出", "excel", "Excel", "表格", "一批", "一份", "清单", "一些", "几个", "的", "看看", "介绍"
]

# SIGNAL_RULES 已废弃
SIGNAL_RULES = []

DEFAULT_SCORE_MIN = 11
DEFAULT_LIMIT = 20


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _contains_any(text_value: str, keywords: Iterable[str]) -> bool:
    return any(keyword and keyword in text_value for keyword in keywords)


def _parse_number(value: Any) -> float:
    """将收入、增长率等混合文本解析为数字。收入统一近似为万元。"""
    raw = _clean_text(value)
    if not raw or raw in {"-", "--", "无", "暂无", "None", "nan"}:
        return 0.0

    normalized = raw.replace(",", "").replace("，", "").replace(" ", "")
    match = re.search(r"-?\d+(?:\.\d+)?", normalized)
    if not match:
        return 0.0

    number = float(match.group())
    if "亿" in normalized:
        number *= 10000
    return number


def _parse_percent(value: Any) -> float:
    raw = _clean_text(value).replace("%", "").replace("％", "")
    match = re.search(r"-?\d+(?:\.\d+)?", raw)
    return float(match.group()) if match else 0.0


def extract_days_limit(k: str) -> int:
    import re
    if re.search(r'(今年|本年)', k):
        now = datetime.now()
        return (now - datetime(now.year, 1, 1)).days or 1
    if re.search(r'(半年|6个月|六个月)', k):
        return 180
    if re.search(r'(一个季度|1个季度|三个月|3个月)', k):
        return 90
    if re.search(r'(一个月|1个月|30天)', k):
        return 30
    if re.search(r'(一周|一星期|7天|七天)', k):
        return 7
    if re.search(r'(全部|所有时间|不限时间)', k):
        return 3650
    return 30

def extract_limit(k: str) -> Optional[int]:
    import re
    match = re.search(r'(?:推荐|前|查|找|给|展示|列出)?\s*(\d+|[一二两三四五六七八九十百千万]+)\s*(?:个|家|名|条|份)', k)
    if match:
        num_str = match.group(1)
        if num_str.isdigit():
            return int(num_str)
        
        num_map = {'零': 0, '一': 1, '二': 2, '两': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9}
        unit_map = {'十': 10, '百': 100, '千': 1000, '万': 10000}
        
        if num_str in num_map:
            return num_map[num_str]
        if num_str == '十':
            return 10
            
        result = 0
        tmp = 0
        for char in num_str:
            if char in num_map:
                tmp = num_map[char]
            elif char in unit_map:
                unit = unit_map[char]
                if tmp == 0 and unit == 10:
                    tmp = 1
                result += tmp * unit
                tmp = 0
        result += tmp
        return result
    return None

def parse_filters(keyword: Optional[str], score_min: int = DEFAULT_SCORE_MIN, raw_text: str = None) -> Dict[str, Any]:
    text_value = _clean_text(raw_text if raw_text else keyword)
    district = None
    industry = None

    for alias, full_name in DISTRICT_ALIASES.items():
        if alias and alias in text_value:
            district = full_name
            break

    for ind in INDUSTRY_KEYWORDS:
        if ind and ind.lower() in text_value.lower():
            industry = "人工智能" if ind == "AI" else ind
            break

    days_limit = extract_days_limit(text_value)
    limit = extract_limit(text_value)

    cleaned = text_value
    # 移除包含数量的表达方式
    import re
    cleaned = re.sub(r'(?:推荐|前|查|找|给|展示|列出)?\s*(\d+|[一二两三四五六七八九十百千万]+)\s*(?:个|家|名|条|份)', ' ', cleaned)
    
    for word in NOISE_WORDS:
        cleaned = cleaned.replace(word, " ")
    if district:
        cleaned = cleaned.replace(district, " ").replace(district.replace("新区", "").replace("区", ""), " ")
    if industry:
        cleaned = cleaned.replace(industry, " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()

    return {
        "district": district,
        "industry": industry,
        "keyword": cleaned or None,
        "score_min": score_min,
        "days_limit": days_limit,
        "limit": limit,
    }


def _build_like_filter(field_exprs: List[str], param_name: str) -> str:
    return "(" + " OR ".join([f"{field} LIKE :{param_name}" for field in field_exprs]) + ")"


def fetch_candidate_enterprises(filters: Dict[str, Any], limit: int = 300) -> List[Dict[str, Any]]:
    import utils.db_helper as db
    import json
    district = filters.get("district")
    district_param = None if district == "上海市" else district
    days_limit = filters.get("days_limit", 30)
    
    # We fetch more than limit because we will filter in python
    raw_data = db.fetch_weixin_extract_data(limit=1000, district=district_param, days_limit=days_limit)
    
    keyword = filters.get("keyword") or ""
    industry = filters.get("industry") or ""
    
    keyword_lower = keyword.lower()
    industry_lower = industry.lower()
    
    results = []
    for row in raw_data:
        ent_name = row.get("ent_name") or ""
        ent_name = ent_name.strip()
        
        # 过滤掉大模型提取失败或无有效名称的占位符
        if not ent_name or ent_name in {"-", "无", "未知", "不适用", "NA"} or len(ent_name) < 2:
            continue
            
        ind = row.get("industry") or ""
        title = row.get("title") or ""
        
        if industry_lower and industry_lower not in ind.lower() and industry_lower not in ent_name.lower():
            continue
            
        if keyword_lower and keyword_lower not in ent_name.lower() and keyword_lower not in title.lower() and keyword_lower not in ind.lower():
            continue
            
        # Map to the format expected by the rest of potential.py
        results.append({
            "name": ent_name,
            "short_name": "", 
            "customer_name": ent_name,
            "region": row.get("district"),
            "province": "上海市",
            "city": "上海市",
            "industry": ind,
            "industry_alt": "",
            "marketing_industry_l1": "",
            "group_industry_l1": "",
            "ranking_name": "",
            "ranking_type": "",
            "qualification": "",
            "revenue_2024": "",
            "revenue_2025": "",
            "growth_rate": "",
            "subsidy_amount": "",
            "subsidy_rule": "",
            "account_manager": "",
            "account_manager_department": "",
            "ranking_link": "",
            "is_new_hope_customer": "",
            "registered_at": "",
            "regional_score": row.get("score") or 0,
            "signals_json": json.dumps((row.get("tags") or []) + (row.get("hit") or []), ensure_ascii=False),
            "latest_title": title,
            "latest_date": row.get("release_time_raw"),
            "latest_link": row.get("link"),
        })
        if len(results) >= limit:
            break

    company_names = [r["name"] for r in results if r.get("name")]
    if company_names:
        import datetime
        now = datetime.datetime.now()
        placeholders = ", ".join([f":name_{i}" for i in range(len(company_names))])
        params = {f"name_{i}": name for i, name in enumerate(company_names)}
        
        # 1. 关联查询标准企业全称 (brief2full_name)
        sql_brief = f"SELECT `brief_name`, `full_name` FROM brief2full_name WHERE `brief_name` IN ({placeholders})"
        try:
            brief_records = db.query_business_db(sql_brief, params)
            brief_map = {r.get("brief_name"): r.get("full_name") for r in brief_records if r.get("brief_name") and r.get("full_name")}
            for row in results:
                ent_name = row["name"]
                if ent_name in brief_map:
                    row["customer_name"] = brief_map[ent_name]
        except Exception as e:
            db.log_event(None, "potential", "WARNING", f"关联企业标准名称失败: {e}")
        
        # 2. 关联查询榜单信息 (ranking_ent_dtl_clue)
        sql = f"SELECT `企业名称`, `资质名称`, `榜单名称`, `链接`, `到期时间`, `榜单废弃` FROM ranking_ent_dtl_clue WHERE `企业名称` IN ({placeholders})"
        try:
            rankings = db.query_business_db(sql, params)
            ranking_map = {}
            for r in rankings:
                discarded = str(r.get("榜单废弃") or "").strip()
                if discarded in ["1", "是", "true", "True", "废弃"]:
                    continue
                
                expire_str = str(r.get("到期时间") or "").strip()
                if expire_str and expire_str not in ["-", "无", "0", "None"]:
                    try:
                        expire_str = expire_str.replace("/", "-")
                        parts = expire_str.split()[0].split("-")
                        if len(parts) >= 3:
                            expire_date = datetime.datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                            if expire_date < now:
                                continue
                    except Exception:
                        pass
                
                ent_name = r.get("企业名称")
                if not ent_name:
                    continue
                if ent_name not in ranking_map:
                    ranking_map[ent_name] = {"qualifications": [], "ranking_names": [], "ranking_links": []}
                    
                qual = str(r.get("资质名称") or "").strip()
                if qual and qual not in ranking_map[ent_name]["qualifications"]:
                    ranking_map[ent_name]["qualifications"].append(qual)
                    
                rank_name = str(r.get("榜单名称") or "").strip()
                if rank_name and rank_name not in ranking_map[ent_name]["ranking_names"]:
                    ranking_map[ent_name]["ranking_names"].append(rank_name)
                    
                link = str(r.get("链接") or "").strip()
                if link and link not in ranking_map[ent_name]["ranking_links"]:
                    ranking_map[ent_name]["ranking_links"].append(link)
            
            for row in results:
                ent_name = row["name"]
                if ent_name in ranking_map:
                    row["qualification"] = "，".join(ranking_map[ent_name]["qualifications"])
                    row["ranking_name"] = "，".join(ranking_map[ent_name]["ranking_names"])
                    row["ranking_link"] = "，".join(ranking_map[ent_name]["ranking_links"])
        except Exception as e:
            db.log_event(None, "potential", "WARNING", f"关联榜单信息失败: {e}")
            
    return results


def fetch_enterprise_signals(candidates: List[Dict[str, Any]], max_candidates: int = 100) -> Dict[str, Dict[str, Any]]:
    return {}



def score_enterprise(row: Dict[str, Any]) -> Tuple[int, List[str], Dict[str, int]]:
    import json
    score_parts: Dict[str, int] = {}
    tags: List[str] = []

    regional_score = int(row.get("regional_score") or 0)
    score_parts["商机基础分"] = regional_score

    signals_json = row.get("signals_json")
    if signals_json:
        try:
            signal_tags = json.loads(signals_json)
            if isinstance(signal_tags, list):
                tags.extend(signal_tags)
        except:
            pass

    return regional_score, list(dict.fromkeys(tags)), score_parts


def build_reason(row: Dict[str, Any], tags: List[str]) -> str:
    tag_text = "、".join(tags[:4]) if tags else "优质企业"
    latest_title = _clean_text(row.get("latest_title"))
    if latest_title:
        return f"企业具备{tag_text}等特征，近期动态“{latest_title}”显示其业务扩张或投入意愿较强，适合优先跟进。"
    return f"企业具备{tag_text}等特征，符合高潜客户筛选标准，建议纳入重点客户池持续跟进。"


def build_next_action(row: Dict[str, Any], tags: List[str]) -> str:
    industry = " ".join([
        _clean_text(row.get("industry")),
        _clean_text(row.get("industry_alt")),
        _clean_text(row.get("marketing_industry_l1")),
        _clean_text(row.get("group_industry_l1")),
    ])
    if _contains_any(industry + " ".join(tags), ["人工智能", "算力", "大数据", "云计算", "软件", "信息技术"]):
        return "建议客户经理优先跟进云资源、专线、算力和数据服务需求。"
    if _contains_any(industry + " ".join(tags), ["智能制造", "机器人", "汽车", "新能源", "新材料", "半导体", "集成电路"]):
        return "建议优先跟进工业互联网、园区专线、边缘计算和 ICT 集成需求。"
    return "建议客户经理开展首次触达，确认专线、云资源、ICT 集成和政策申报需求。"


def _build_enterprise_rag_document(row: Dict[str, Any]) -> Dict[str, Any]:
    name = _clean_text(row.get("name"))
    region = _clean_text(row.get("region")) or _clean_text(row.get("city")) or _clean_text(row.get("province"))
    industry = _clean_text(row.get("industry"))
    import json
    try:
        signals = json.loads(row.get("signals_json") or "[]")
    except:
        signals = []
    content = "\\n".join([
        f"企业名称：{name}",
        f"行政区：{region}",
        f"行业：{industry}",
        f"近期新闻信号：{'、'.join(signals)}",
        f"最新动态标题：{_clean_text(row.get('latest_title'))}",
    ])
    return {
        "title": name or "高潜候选企业",
        "content": content,
        "publish_date": _clean_text(row.get("latest_date")) or _clean_text(row.get("registered_at")) or datetime.now().strftime("%Y-%m-%d"),
        "source": "weixin_deepseek_extract_d",
        "link": _clean_text(row.get("latest_link")),
        "company": name,
        "district": region,
        "industry": industry,
        "doc_type": "enterprise_profile",
    }


def _build_news_rag_documents(candidates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    documents = []
    for row in candidates[:100]:
        name = _clean_text(row.get("name"))
        latest_title = _clean_text(row.get("latest_title"))
        if not name or not latest_title:
            continue
        try:
            import json
            signals = json.loads(row.get("signals_json") or "[]")
        except:
            signals = []
        content = "\\n".join([
            f"企业名称：{name}",
            f"新闻标题：{latest_title}",
            f"命中信号：{'、'.join(signals)}",
            f"新闻信号得分：{_clean_text(row.get('regional_score'))}",
        ])
        documents.append({
            "title": latest_title,
            "content": content,
            "publish_date": _clean_text(row.get("latest_date")),
            "source": "opportunity_articles",
            "link": _clean_text(row.get("latest_link")),
            "company": name,
            "district": _clean_text(row.get("region")) or _clean_text(row.get("city")) or _clean_text(row.get("province")),
            "industry": _clean_text(row.get("industry")) or _clean_text(row.get("industry_alt")),
            "doc_type": "news_signal",
        })
    return documents


def _fetch_policy_rag_documents(filters: Dict[str, Any], limit: int = 12) -> List[Dict[str, Any]]:
    keywords = [filters.get("industry"), filters.get("district"), filters.get("keyword")]
    keywords = [item for item in keywords if item]
    if not keywords:
        return []

    params = {f"kw{idx}": f"%{kw}%" for idx, kw in enumerate(keywords)}
    like_parts = []
    for idx in range(len(keywords)):
        like_parts.append(f"`标题` LIKE :kw{idx} OR `正文` LIKE :kw{idx} OR `关键词` LIKE :kw{idx}")
    sql = text(f"""
        SELECT `标题` AS title, `正文` AS content, `网址` AS link, `发布单位` AS source, `发布时间` AS publish_date
        FROM zq_dtl_onenet_all
        WHERE {' OR '.join(like_parts)}
        ORDER BY `发布时间` DESC
        LIMIT {int(limit)}
    """)

    try:
        rows = db.query_business_db(str(sql), params)
    except Exception:
        return []

    documents = []
    for row in rows:
        documents.append({
            "title": row.get("title") or "政策信号",
            "content": row.get("content") or "",
            "publish_date": row.get("publish_date") or "",
            "source": row.get("source") or "政策库",
            "link": row.get("link") or "",
            "company": "policy",
            "district": filters.get("district") or "",
            "industry": filters.get("industry") or "",
            "doc_type": "policy_signal",
        })
    return documents


def _build_potential_rag_query(filters: Dict[str, Any]) -> str:
    return "\n".join([
        f"区域：{filters.get('district') or '不限'}",
        f"行业：{filters.get('industry') or '不限'}",
        f"关键词：{filters.get('keyword') or '高潜客户'}",
        "任务：筛选高潜客户，关注收入增长、榜单资质、政策补贴、融资上市、重大签约、扩产落地、技术突破、政府关注和数字化业务需求。",
    ])


def _aggregate_rag_by_company(retrieved_docs: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}
    for doc in retrieved_docs:
        metadata = doc.get("metadata") or {}
        company = metadata.get("company") or ""
        if not company or company == "policy":
            continue
        score = float(doc.get("rerank_score", doc.get("final_score", 0.0)) or 0.0)
        item = result.setdefault(company, {"best_score": 0.0, "evidence": []})
        item["best_score"] = max(item["best_score"], score)
        if len(item["evidence"]) < 3:
            item["evidence"].append({
                "title": metadata.get("title") or "RAG证据",
                "source": metadata.get("source") or "",
                "link": metadata.get("link") or "",
                "doc_type": metadata.get("doc_type") or "",
                "score": round(score, 4),
            })
    return result


def _rag_bonus(score: float) -> int:
    if score >= 0.85:
        return 10
    if score >= 0.70:
        return 7
    if score >= 0.55:
        return 5
    if score >= 0.40:
        return 2
    return 0


def _run_potential_rag(filters: Dict[str, Any], candidates: List[Dict[str, Any]], user_id: int = None) -> Dict[str, Dict[str, Any]]:
    documents = []
    for row in candidates[:300]:
        name = _clean_text(row.get("name"))
        if name:
            documents.append(_build_enterprise_rag_document(row))
    documents.extend(_build_news_rag_documents(candidates))
    documents.extend(_fetch_policy_rag_documents(filters))

    if not documents:
        return {}

    api_key = os.getenv("DEEPSEEK_API_KEY")
    base_url = os.getenv("DEEPSEEK_API_BASE", "https://api.deepseek.com")
    model_name = os.getenv("OPENAI_MODEL_NAME", "deepseek-chat")
    query = _build_potential_rag_query(filters)

    try:
        rag = RAGEngine(documents=documents)
        retrieved_docs = rag.retrieve(query, top_k=40)
        if api_key and "your_api_key" not in api_key:
            retrieved_docs = rag.rerank(query, retrieved_docs, api_key, base_url, model_name)
        else:
            for doc in retrieved_docs:
                doc["rerank_score"] = doc.get("final_score", 0.0)
        company_rag = _aggregate_rag_by_company(retrieved_docs)
        db.log_event(user_id, "potential", "INFO", f"高潜客户 RAG 检索完成，命中企业数: {len(company_rag)}")
        return company_rag
    except Exception as exc:
        db.log_event(user_id, "potential", "WARNING", f"高潜客户 RAG 检索失败，降级规则评分: {exc}")
        return {}

    api_key = os.getenv("DEEPSEEK_API_KEY")
    base_url = os.getenv("DEEPSEEK_API_BASE", "https://api.deepseek.com")
    model_name = os.getenv("OPENAI_MODEL_NAME", "deepseek-chat")
    query = _build_potential_rag_query(filters)

    try:
        rag = RAGEngine(documents=documents)
        retrieved_docs = rag.retrieve(query, top_k=40)
        if api_key and "your_api_key" not in api_key:
            retrieved_docs = rag.rerank(query, retrieved_docs, api_key, base_url, model_name)
        else:
            for doc in retrieved_docs:
                doc["rerank_score"] = doc.get("final_score", 0.0)
        company_rag = _aggregate_rag_by_company(retrieved_docs)
        db.log_event(user_id, "potential", "INFO", f"高潜客户 RAG 检索完成，命中企业数: {len(company_rag)}")
        return company_rag
    except Exception as exc:
        db.log_event(user_id, "potential", "WARNING", f"高潜客户 RAG 检索失败，降级规则评分: {exc}")
        return {}


def _apply_rag_to_candidate(item: Dict[str, Any], rag_info: Dict[str, Any]) -> Dict[str, Any]:
    regional_score = item.get("score", 0)
    
    if not rag_info:
        item["rag_score"] = 0.0
        item["rag_evidence"] = []
        item["score"] = regional_score
        item["level"] = "HOT" if regional_score >= 60 else "关注"
        return item

    score = float(rag_info.get("best_score") or 0.0)
    evidence = rag_info.get("evidence") or []
    item["rag_score"] = round(score, 4)
    item["rag_evidence"] = evidence
    
    item["score"] = regional_score
    item["level"] = "HOT" if regional_score >= 60 else "关注"
    
    if evidence:
        signals = item.setdefault("signals", [])
        if "RAG强相关证据" not in signals:
            signals.append("RAG强相关证据")
        title = evidence[0].get("title") or "相关证据"
        item["reason"] = f"{item.get('reason', '')} RAG 证据“{title}”进一步证明其与当前筛选目标相关，建议优先核实业务需求。"
    return item


def _normalize_candidate(row: Dict[str, Any], score: int, tags: List[str], score_parts: Dict[str, int]) -> Dict[str, Any]:
    region = _clean_text(row.get("region")) or _clean_text(row.get("city")) or _clean_text(row.get("province"))
    industry = _clean_text(row.get("industry")) or "未标注"
    level = "HOT" if score >= 60 else "关注"
    return {
        "name": _clean_text(row.get("name")),
        "short_name": _clean_text(row.get("short_name")),
        "customer_name": _clean_text(row.get("customer_name")) or _clean_text(row.get("name")),
        "score": score,
        "level": level,
        "industry": industry,
        "industry_alt": _clean_text(row.get("industry_alt")),
        "marketing_industry_l1": _clean_text(row.get("marketing_industry_l1")),
        "region": region or "未标注",
        "signals": tags[:6],
        "reason": build_reason(row, tags),
        "next_action": build_next_action(row, tags),
        "account_manager": "",
        "qualification": _clean_text(row.get("qualification")),
        "revenue_2024": "",
        "growth_rate": "",
        "subsidy_amount": "",
        "subsidy_rule": "",
        "ranking_name": _clean_text(row.get("ranking_name")),
        "ranking_type": "",
        "ranking_link": _clean_text(row.get("ranking_link")),
        "is_new_hope_customer": "",
        "latest_title": _clean_text(row.get("latest_title")),
        "latest_date": _clean_text(row.get("latest_date")),
        "link": _clean_text(row.get("latest_link")),
        "latest_content": _clean_text(row.get("latest_content")),
        "score_parts": score_parts,
    }


def _fallback_from_mock(filters: Dict[str, Any], limit: int = DEFAULT_LIMIT) -> List[Dict[str, Any]]:
    keyword = filters.get("keyword") or filters.get("industry") or ""
    items = []
    for client in POTENTIAL_CLIENTS:
        if keyword and keyword.lower() not in (client.get("name", "") + client.get("industry", "")).lower():
            continue
        score = int(client.get("score") or 0)
        items.append({
            "name": client.get("name"),
            "short_name": "",
            "score": score,
            "level": "HOT" if score >= 60 else "关注",
            "industry": client.get("industry"),
            "region": filters.get("district") or "未标注",
            "signals": ["Mock兜底", "优质线索"],
            "reason": client.get("reason"),
            "next_action": "建议客户经理优先确认专线、云资源和 ICT 集成需求。",
            "account_manager": client.get("contact", "待分配"),
            "qualification": "",
            "revenue_2024": client.get("scale", ""),
            "growth_rate": "",
            "subsidy_rule": "",
            "latest_title": "",
            "latest_date": "",
            "link": client.get("link", ""),
            "score_parts": {},
        })
    return items[:limit] or []


def get_recommendations(filters: Dict[str, Any], limit: int = DEFAULT_LIMIT, skip_rag: bool = False) -> Tuple[List[Dict[str, Any]], int]:
    try:
        candidates = fetch_candidate_enterprises(filters)
        company_rag = {} if skip_rag else _run_potential_rag(filters, candidates)
        recommendations = []
        ranked_all = []
        score_min = int(filters.get("score_min") or DEFAULT_SCORE_MIN)
        seen_names = set()

        for row in candidates:
            name = _clean_text(row.get("name"))
            if not name or name in seen_names:
                continue
            
            if not _clean_text(row.get("latest_link")):
                continue

            seen_names.add(name)
            score, tags, score_parts = score_enterprise(row)
            normalized = _normalize_candidate(row, score, tags, score_parts)
            normalized = _apply_rag_to_candidate(normalized, company_rag.get(name, {}))
            if int(normalized.get("score") or 0) >= score_min:
                ranked_all.append(normalized)
                recommendations.append(normalized)

        if skip_rag:
            ranked_all.sort(key=lambda item: (item["score"], item.get("rag_score", 0.0)), reverse=True)
            return ranked_all[:limit], len(ranked_all)

        recommendations.sort(key=lambda item: (item["score"], item.get("rag_score", 0.0)), reverse=True)
        if recommendations:
            return recommendations[:limit], len(recommendations)

        # 行政区类查询如果没有超过阈值的企业，不直接返回空；保底返回该区域综合得分最高的若干企业。
        ranked_all.sort(key=lambda item: (item["score"], item.get("rag_score", 0.0)), reverse=True)
        if ranked_all:
            return ranked_all[: min(limit, 10)], len(ranked_all)
        return [], 0
    except Exception as exc:
        import traceback
        import utils.db_helper as db
        db.log_event(None, "potential", "ERROR", f"高潜客户数据库检索失败，启用兜底数据: {exc}", traceback.format_exc())
        mock_items = _fallback_from_mock(filters, limit=limit)
        return mock_items, len(mock_items)



def write_items_to_excel(items: List[Dict[str, Any]], safe_scope: str) -> str:
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    static_dir = os.path.join(base_dir, "static", "generated")
    ensure_dir(static_dir)

    safe_scope = re.sub(r"[\\/:*?\"<>|\s]+", "_", safe_scope)[:30]
    filename = f"high_potential_{safe_scope}_{datetime.now().strftime('%Y%m%d')}_{uuid.uuid4().hex[:6]}.xlsx"
    excel_path = os.path.join(static_dir, filename)

    if items:
        df = pd.DataFrame(items)
        df_excel = df.rename(columns={
            "name": "企业名称",
            "customer_name": "企业标准名称",
            "region": "行政区",
            "industry": "重点行业一",
            "industry_alt": "重点行业二",
            "marketing_industry_l1": "重点行业三",
            "score": "推荐评分",
            "level": "推荐等级",
            "signals": "命中信号",
            "reason": "推荐理由",
            "next_action": "下一步动作",
            "qualification": "资质名称",
            "ranking_name": "榜单名称",
            "latest_title": "最新动态标题",
            "latest_date": "最新动态日期",
            "link": "新闻原文链接",
            "ranking_link": "榜单原文链接",
        })
        if "命中信号" in df_excel.columns:
            df_excel["命中信号"] = df_excel["命中信号"].apply(lambda value: "、".join(value) if isinstance(value, list) else value)
        keep_columns = [
            "企业名称", "企业标准名称", "行政区", "重点行业一", "重点行业二", "重点行业三",
            "推荐评分", "推荐等级", "命中信号", "推荐理由", "下一步动作",
            "资质名称", "榜单名称", "最新动态标题", "最新动态日期", "新闻原文链接", "榜单原文链接",
        ]
        df_excel = df_excel[[column for column in keep_columns if column in df_excel.columns]]
    else:
        df_excel = pd.DataFrame([{"提示": "暂无符合条件的高潜客户"}])

    df_excel.to_excel(excel_path, index=False)
    _apply_excel_hyperlinks(excel_path, ["新闻原文链接", "榜单原文链接"])
    return excel_path


def _build_export_url(filters: Dict[str, Any]) -> str:
    from urllib.parse import urlencode
    params = {"score_min": filters.get("score_min") or DEFAULT_SCORE_MIN}
    if filters.get("district"):
        params["district"] = filters["district"]
    if filters.get("industry"):
        params["industry"] = filters["industry"]
    if filters.get("keyword"):
        params["keyword"] = filters["keyword"]
    if filters.get("limit"):
        params["limit"] = filters["limit"]
    return "/api/potential/export?" + urlencode(params)


def handle(keyword: str, user_id: int = None, raw_text: str = None) -> dict:
    if not keyword and not raw_text:
        return {
            "type": "text",
            "content": "请问您想挖掘哪个行政区或哪个行业的高潜客户？（例如：浦东新区、人工智能行业等）"
        }
    filters = parse_filters(keyword, raw_text=raw_text)
    db.log_event(user_id, "potential", "INFO", f"开始检索高潜客户线索。过滤条件: {filters}")

    user_limit = filters.get("limit")
    limit = user_limit or DEFAULT_LIMIT
    items, total_count = get_recommendations(filters, limit=limit)
    if user_limit:
        total_count = len(items)
    
    if items:
        try:
            items = asyncio.run(_async_generate_all_reasons(items))
        except Exception as e:
            db.log_event(user_id, "potential", "WARNING", f"大模型润色推荐理由失败: {e}")
            
    district_label = filters.get("district") or "全市"
    industry_label = filters.get("industry")
    title_parts = [district_label]
    if industry_label:
        title_parts.append(industry_label)
    title = "".join(title_parts) + "高潜客户推荐"

    if items:
        strong_count = sum(1 for item in items if int(item.get("score") or 0) >= int(filters.get("score_min") or DEFAULT_SCORE_MIN))
        if strong_count:
            summary = f"满足过滤条件的高潜企业共有 {total_count} 家。为保证质量，已为您精选出排名前 {len(items)} 家进行展示。如需查看完整清单，请点击下方按钮导出完整版 Excel 表格。"
        else:
            summary = f"暂未发现评分严格达标的企业，已为您放宽条件，展示 {len(items)} 家候选企业供参考。您可以点击下方按钮导出更多候选结果。"
    else:
        summary = "暂未筛选到符合条件的高潜客户，建议放宽区域、行业或评分条件后重试。"

    actions = []
    if items:
        excel_url = _build_export_url(filters)
        actions.append({
            "label": f"导出所有满足条件的高潜客户 ({total_count}家)",
            "type": "download",
            "url": excel_url,
        })

    payload = {
        "type": "high_potential_customers",
        "title": title,
        "summary": summary,
        "filters": filters,
        "items": items,
        "actions": actions,
        "count": len(items),
    }
    db.log_event(user_id, "potential", "INFO", f"高潜客户推荐生成完成，数量: {len(items)}")
    return payload


def _apply_excel_hyperlinks(excel_path: str, link_columns: List[str]) -> None:
    """将指定列中的 URL 文本转换为 Excel 可点击超链接。"""
    try:
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        header_map = {
            str(cell.value or "").strip(): idx
            for idx, cell in enumerate(sheet[1], start=1)
        }
        for column_name in link_columns:
            column_idx = header_map.get(column_name)
            if not column_idx:
                continue
            for row_idx in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=column_idx)
                url = _clean_text(cell.value)
                if not (url.startswith("http://") or url.startswith("https://")):
                    continue
                cell.hyperlink = url
                cell.value = "查看原文"
                cell.style = "Hyperlink"
                cell.font = Font(color="0563C1", underline="single")
        workbook.save(excel_path)
    except Exception as exc:
        db.log_event(None, "potential", "WARNING", f"Excel 超链接格式化失败，保留原始 URL 文本: {exc}")


def export_excel(
    district: Optional[str] = None,
    industry: Optional[str] = None,
    keyword: Optional[str] = None,
    score_min: int = DEFAULT_SCORE_MIN,
    limit: Optional[int] = None,
    user_id: Optional[int] = None,
) -> str:
    filter_text = " ".join([value for value in [district, industry, keyword] if value])
    filters = parse_filters(filter_text, score_min=score_min)
    if district:
        filters["district"] = district
    if industry:
        filters["industry"] = industry
    if keyword:
        filters["keyword"] = keyword
    filters["score_min"] = score_min
    if limit is not None:
        filters["limit"] = limit

    export_limit = limit if limit is not None else 500
    items, _ = get_recommendations(filters, limit=export_limit, skip_rag=True)

    safe_scope = district or industry or keyword or "全市"
    excel_path = write_items_to_excel(items, safe_scope)
    db.log_event(user_id, "potential", "INFO", f"高潜客户 Excel 导出成功: {excel_path}")
    return excel_path


async def _async_generate_reason_with_client(client: AsyncOpenAI, model_name: str, item: dict) -> dict:
    rag_evidences = [e.get("title") for e in item.get("rag_evidence", [])]
    
    prompt = f"""
请作为顶级的大客户经理和售前架构师，为高潜企业“{item.get('name')}”撰写一段个性化的推荐理由和下一步行动建议。
【已知企业信息】：
行业：{item.get('industry', '未知')}
已知资质：{item.get('qualification', '无')}
收入增速：{item.get('growth_rate', '未知')}
命中信号标签：{','.join(item.get('signals', []))}
最新动态标题：{item.get('latest_title', '无')}
最新动态内容（原文摘录）：{item.get('latest_content', '无')}
RAG补充证据：{rag_evidences}

【撰写要求】：
1. 深入研读新闻“原文摘录”，提炼出能真正体现其扩张、转型、建厂、出海、技术升级等潜在 IT/通信/算力 需求的核心情报。
2. 推荐理由（reason）：一到两句大白话分析。**必须彻底摆脱模板化套话（严禁使用“涉及大规模数字化展示和运营”、“必然需要高可靠网络、云资源和智能系统集成，潜在ICT需求巨大”这类万金油句式）**。请直接指出新闻里发生了什么具体的业务事件，这会导致什么样的具体采购需求。
3. 下一步动作（next_action）：给出一句话极其具体的销售跟进建议。**严禁使用“提供整体ICT集成服务”、“跟进云资源方案”这类空泛套话**。如果是新建园区，就写建议切入弱电与专线；如果是出海，就推跨境专网；如果是新签AI战略，就推算力租赁等，必须要跟新闻里的动作挂钩。
4. 严格以 JSON 格式返回，包含 'reason' 和 'next_action' 两个字符串字段，不要有 ```json 标记。
"""
    try:
        response = await client.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": "你是一个顶级售前架构师，擅长用最敏锐的商业视角深度剖析企业新闻与资质背后的商机。"},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            temperature=0.3,
            timeout=15.0
        )
        # 兼容处理大模型返回的空字符或非合法 JSON 避免 Expecting value 报错
        content = (response.choices[0].message.content or "").strip()
        if content.startswith("```json"):
            content = content[7:-3].strip()
        if content:
            res = json.loads(content)
            if "reason" in res and res["reason"]:
                item["reason"] = res["reason"]
            if "next_action" in res and res["next_action"]:
                item["next_action"] = res["next_action"]
    except Exception as e:
        print(f"Generate reason error for {item.get('name')}: {e}")
    return item

async def _async_generate_all_reasons(items: list) -> list:
    api_key = os.getenv("DEEPSEEK_API_KEY")
    base_url = os.getenv("DEEPSEEK_API_BASE", "https://api.deepseek.com")
    model_name = os.getenv("OPENAI_MODEL_NAME", "deepseek-chat")
    if not api_key or "your_api_key" in api_key:
        return items
        
    client = AsyncOpenAI(api_key=api_key, base_url=base_url)
    try:
        tasks = [_async_generate_reason_with_client(client, model_name, item) for item in items]
        await asyncio.gather(*tasks)
    finally:
        await client.close()
    return items

